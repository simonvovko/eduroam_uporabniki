# -*- encoding: utf-8 -*-
'''
pip install openpyxl

Delo z Excelom:
https://automatetheboringstuff.com/chapter12/
'''

import openpyxl
from random import randint
import string
import datetime

def popraviEMSO(emso):
    emso = emso.strip()
    return (emso[:-3] + str(int(emso[10:])+1).zfill(3))

def generirajGeslo(d):
    geslo = ''

    for i in range(d):
        g = randint(1,3)
        
        if g == 1:
            geslo += chr(randint(65, 90))
        elif g == 2:
            geslo += chr(randint(97, 122))
        elif g == 3:
            geslo += chr(randint(48, 57))

    return geslo

def sestaviUporabniskoIme(ime, priimek):
    ime     = ime.strip()
    priimek = priimek.strip()
    
    dijak = ime + ' ' + priimek
    dijak = dijak.lower()

    dijak = dijak.replace('č', 'c')
    dijak = dijak.replace('š', 's')
    dijak = dijak.replace('ž', 'z')
    dijak = dijak.replace('ć', 'c')
    dijak = dijak.replace(' ', '.')
    dijak = dijak.replace('\t', '.')

    return dijak

def urediDatum(datum):
    return datum.replace(" ", "")

def getOddExp(razred):
    razred = razred.upper()
    # Tehniki računalništva
    if razred in ('R1A', 'R1B', 'R1C'): return '1R', 4
    elif razred in ('R2A', 'R2B', 'R2C'): return '2R', 3
    elif razred in ('R3A', 'R3B', 'R3C'): return '3R', 2
    elif razred in ('R4A', 'R4B', 'R4C'): return '4R', 1

    # Tehniška gimnazija
    elif razred in ('T1A', 'T1B', 'T1C'): return '1TG', 4
    elif razred in ('T2A', 'T2B', 'T2C'): return '2TG', 3
    elif razred in ('T3A', 'T3B', 'T3C'): return '3TG', 2
    elif razred in ('T4A', 'T4B', 'T4C'): return '4TG', 1

    # Elektrotehniki
    elif razred in ('E1A', 'E1B'): return '1E', 4
    elif razred in ('E2A', 'E2B'): return '2E', 3
    elif razred in ('E3A', 'E3B'): return '3E', 2
    elif razred in ('E4A', 'E4B'): return '4E', 1

    # PTI - Rač
    elif razred in ('E1TA', 'E1TB'): return '1Ep', 2
    elif razred in ('E2TA', 'E2TB'): return '2Ep', 1

    # PTI - Ele
    elif razred in ('R1TA', 'R1TB'): return '1Rp', 2
    elif razred in ('R2TA', 'R2TB'): return '2Rp', 1

    # Elektrikarji
    elif razred in ('E1C'): return '1E3', 3
    elif razred in ('E2C'): return '2E3', 2
    elif razred in ('E3C'): return '3E3', 1

    # Računalnikarji
    elif razred in ('R1E'): return '1R3', 3
    elif razred in ('R2E'): return '2R3', 2
    elif razred in ('R3E'): return '3R3', 1

    # ????
    else: return '????', -1

# odprem datoteko, ki jo izvozim iz eAsistenta
dijaki = openpyxl.load_workbook('ime_datoteke_z_easistenta.xlsx')

uvoz = openpyxl.load_workbook('uvoz.xlsx')
u_sheet = uvoz.active

stDijakov = 0

# preberi leto, ki ga potrebujem, za nastavitev trajanja uporabniškega imena
now = datetime.datetime.now()
expyr = now.year

# grem čez vse sheet-e (vsak sheet hrani podatke o dijakih v oddelku)
for sheet in dijaki:
    # izberem sheet
    print("Obdelujem", sheet.title, end=': ')

    # grem čez vse vrstice (kolikor jih je)
    for row in range(3, sheet.max_row + 1): # podatki v izvozu se začno s tretjo vrstico
        print('.', end='')

        ime             = sheet.cell(row=row, column=1).value
        priimek         = sheet.cell(row=row, column=2).value
        datumRojstva    = sheet.cell(row=row, column=3).value
        spol            = sheet.cell(row=row, column=4).value
        emso            = sheet.cell(row=row, column=5).value
        email           = sheet.cell(row=row, column=6).value
        razred          = sheet.cell(row=row, column=7).value

        emso            = popraviEMSO(emso)
        datumRojstva    = urediDatum(datumRojstva)
        geslo           = generirajGeslo(8)
        uporabnisko     = sestaviUporabniskoIme(ime, priimek)

        oddelek, trajanjeSolanja = getOddExp(razred)
        if trajanjeSolanja == -1: expyr = -1

        u_sheet.cell(row=stDijakov+2, column=2).value = ime
        u_sheet.cell(row=stDijakov+2, column=3).value = priimek
        u_sheet.cell(row=stDijakov+2, column=5).value = datumRojstva
        u_sheet.cell(row=stDijakov+2, column=6).value = spol
        u_sheet.cell(row=stDijakov+2, column=7).value = emso
        u_sheet.cell(row=stDijakov+2, column=8).value = 'Šegova ulica'
        u_sheet.cell(row=stDijakov+2, column=9).value = '112'
        u_sheet.cell(row=stDijakov+2, column=10).value = 'Novo mesto'
        u_sheet.cell(row=stDijakov+2, column=11).value = '8000'
        u_sheet.cell(row=stDijakov+2, column=16).value = email
        u_sheet.cell(row=stDijakov+2, column=19).value = oddelek
        u_sheet.cell(row=stDijakov+2, column=23).value = 'RED'
        u_sheet.cell(row=stDijakov+2, column=24).value = 'DIJ'
        u_sheet.cell(row=stDijakov+2, column=32).value = uporabnisko
        u_sheet.cell(row=stDijakov+2, column=33).value = geslo
        u_sheet.cell(row=stDijakov+2, column=34).value = razred
        u_sheet.cell(row=stDijakov+2, column=35).value = expyr + trajanjeSolanja
        
        stDijakov += 1
    print()
        
uvoz.save('uvoz_save.xlsx')
print('Končano')
